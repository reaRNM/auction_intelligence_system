from typing import Dict, List, Optional, Union
import logging
from datetime import datetime, timedelta
import pandas as pd
import plotly.graph_objects as go
from weasyprint import HTML
import openpyxl
from decimal import Decimal

logger = logging.getLogger(__name__)

class BaseReporter:
    """Base class for report generation."""
    
    def __init__(self):
        """Initialize the reporter."""
        self.report_templates = {
            'daily_opportunity': 'templates/reports/daily_opportunity.html',
            'post_auction': 'templates/reports/post_auction.html',
            'tax_prep': 'templates/reports/tax_prep.html'
        }
        
        self.output_formats = ['dashboard', 'pdf', 'excel', 'email']
    
    def generate_report(self, report_type: str, data: Dict, format: str = 'dashboard') -> Union[str, bytes]:
        """Generate a report.
        
        Args:
            report_type: Type of report to generate
            data: Data for the report
            format: Output format
            
        Returns:
            Report in requested format
        """
        try:
            if format not in self.output_formats:
                logger.error(f"Invalid output format: {format}")
                return None
            
            # Generate report based on type
            if report_type == 'daily_opportunity':
                return self._generate_daily_opportunity(data, format)
            elif report_type == 'post_auction':
                return self._generate_post_auction(data, format)
            elif report_type == 'tax_prep':
                return self._generate_tax_prep(data, format)
            else:
                logger.error(f"Invalid report type: {report_type}")
                return None
                
        except Exception as e:
            logger.error(f"Report generation failed: {e}")
            return None
    
    def _generate_daily_opportunity(self, data: Dict, format: str) -> Union[str, bytes]:
        """Generate daily opportunity report.
        
        Args:
            data: Report data
            format: Output format
            
        Returns:
            Report in requested format
        """
        try:
            # Extract data
            opportunities = data.get('opportunities', [])
            risk_reward = data.get('risk_reward', {})
            budget = data.get('budget', {})
            
            if format == 'dashboard':
                return self._create_opportunity_dashboard(opportunities, risk_reward, budget)
            elif format == 'pdf':
                return self._create_opportunity_pdf(opportunities, risk_reward, budget)
            elif format == 'excel':
                return self._create_opportunity_excel(opportunities, risk_reward, budget)
            elif format == 'email':
                return self._create_opportunity_email(opportunities, risk_reward, budget)
                
        except Exception as e:
            logger.error(f"Daily opportunity report generation failed: {e}")
            return None
    
    def _generate_post_auction(self, data: Dict, format: str) -> Union[str, bytes]:
        """Generate post-auction analysis report.
        
        Args:
            data: Report data
            format: Output format
            
        Returns:
            Report in requested format
        """
        try:
            # Extract data
            summary = data.get('summary', {})
            performance = data.get('performance', {})
            patterns = data.get('patterns', {})
            
            if format == 'dashboard':
                return self._create_post_auction_dashboard(summary, performance, patterns)
            elif format == 'pdf':
                return self._create_post_auction_pdf(summary, performance, patterns)
            elif format == 'excel':
                return self._create_post_auction_excel(summary, performance, patterns)
            elif format == 'email':
                return self._create_post_auction_email(summary, performance, patterns)
                
        except Exception as e:
            logger.error(f"Post-auction report generation failed: {e}")
            return None
    
    def _generate_tax_prep(self, data: Dict, format: str) -> Union[str, bytes]:
        """Generate tax preparation report.
        
        Args:
            data: Report data
            format: Output format
            
        Returns:
            Report in requested format
        """
        try:
            # Extract data
            fees = data.get('fees', {})
            inventory = data.get('inventory', {})
            sales_tax = data.get('sales_tax', {})
            
            if format == 'dashboard':
                return self._create_tax_prep_dashboard(fees, inventory, sales_tax)
            elif format == 'pdf':
                return self._create_tax_prep_pdf(fees, inventory, sales_tax)
            elif format == 'excel':
                return self._create_tax_prep_excel(fees, inventory, sales_tax)
            elif format == 'email':
                return self._create_tax_prep_email(fees, inventory, sales_tax)
                
        except Exception as e:
            logger.error(f"Tax preparation report generation failed: {e}")
            return None
    
    def _create_opportunity_dashboard(self, opportunities: List[Dict], risk_reward: Dict, budget: Dict) -> str:
        """Create interactive dashboard for daily opportunities.
        
        Args:
            opportunities: List of opportunities
            risk_reward: Risk/reward matrix
            budget: Budget allocation
            
        Returns:
            HTML dashboard
        """
        try:
            # Create opportunity table
            df = pd.DataFrame(opportunities)
            fig = go.Figure(data=[go.Table(
                header=dict(values=list(df.columns)),
                cells=dict(values=[df[col] for col in df.columns])
            )])
            
            # Create risk/reward scatter plot
            risk_reward_df = pd.DataFrame(risk_reward)
            fig2 = go.Figure(data=[go.Scatter(
                x=risk_reward_df['risk'],
                y=risk_reward_df['reward'],
                mode='markers'
            )])
            
            # Create budget allocation pie chart
            budget_df = pd.DataFrame(budget)
            fig3 = go.Figure(data=[go.Pie(
                labels=budget_df['category'],
                values=budget_df['amount']
            )])
            
            # Combine into dashboard
            dashboard = f"""
            <html>
                <head>
                    <title>Daily Opportunity Report</title>
                    <script src="https://cdn.plot.ly/plotly-latest.min.js"></script>
                </head>
                <body>
                    <h1>Daily Opportunity Report</h1>
                    <div id="opportunities">{fig.to_html()}</div>
                    <div id="risk_reward">{fig2.to_html()}</div>
                    <div id="budget">{fig3.to_html()}</div>
                </body>
            </html>
            """
            
            return dashboard
            
        except Exception as e:
            logger.error(f"Dashboard creation failed: {e}")
            return None
    
    def _create_opportunity_pdf(self, opportunities: List[Dict], risk_reward: Dict, budget: Dict) -> bytes:
        """Create PDF report for daily opportunities.
        
        Args:
            opportunities: List of opportunities
            risk_reward: Risk/reward matrix
            budget: Budget allocation
            
        Returns:
            PDF bytes
        """
        try:
            # Create HTML content
            html_content = f"""
            <html>
                <head>
                    <title>Daily Opportunity Report</title>
                    <style>
                        table {{ border-collapse: collapse; width: 100%; }}
                        th, td {{ border: 1px solid black; padding: 8px; text-align: left; }}
                        th {{ background-color: #f2f2f2; }}
                    </style>
                </head>
                <body>
                    <h1>Daily Opportunity Report</h1>
                    <h2>Top Opportunities</h2>
                    <table>
                        <tr>
                            <th>Item</th>
                            <th>Current Price</th>
                            <th>Estimated Value</th>
                            <th>Potential Profit</th>
                        </tr>
                        {''.join(f"<tr><td>{opp['item']}</td><td>${opp['current_price']}</td><td>${opp['estimated_value']}</td><td>${opp['potential_profit']}</td></tr>" for opp in opportunities)}
                    </table>
                    
                    <h2>Risk/Reward Matrix</h2>
                    <table>
                        <tr>
                            <th>Category</th>
                            <th>Risk Score</th>
                            <th>Reward Score</th>
                        </tr>
                        {''.join(f"<tr><td>{k}</td><td>{v['risk']}</td><td>{v['reward']}</td></tr>" for k, v in risk_reward.items())}
                    </table>
                    
                    <h2>Budget Allocation</h2>
                    <table>
                        <tr>
                            <th>Category</th>
                            <th>Amount</th>
                            <th>Percentage</th>
                        </tr>
                        {''.join(f"<tr><td>{k}</td><td>${v['amount']}</td><td>{v['percentage']}%</td></tr>" for k, v in budget.items())}
                    </table>
                </body>
            </html>
            """
            
            # Convert to PDF
            return HTML(string=html_content).write_pdf()
            
        except Exception as e:
            logger.error(f"PDF creation failed: {e}")
            return None
    
    def _create_opportunity_excel(self, opportunities: List[Dict], risk_reward: Dict, budget: Dict) -> bytes:
        """Create Excel workbook for daily opportunities.
        
        Args:
            opportunities: List of opportunities
            risk_reward: Risk/reward matrix
            budget: Budget allocation
            
        Returns:
            Excel workbook bytes
        """
        try:
            # Create workbook
            wb = openpyxl.Workbook()
            
            # Opportunities sheet
            ws1 = wb.active
            ws1.title = "Opportunities"
            headers = ['Item', 'Current Price', 'Estimated Value', 'Potential Profit']
            ws1.append(headers)
            for opp in opportunities:
                ws1.append([opp['item'], opp['current_price'], opp['estimated_value'], opp['potential_profit']])
            
            # Risk/Reward sheet
            ws2 = wb.create_sheet("Risk/Reward")
            headers = ['Category', 'Risk Score', 'Reward Score']
            ws2.append(headers)
            for category, scores in risk_reward.items():
                ws2.append([category, scores['risk'], scores['reward']])
            
            # Budget sheet
            ws3 = wb.create_sheet("Budget")
            headers = ['Category', 'Amount', 'Percentage']
            ws3.append(headers)
            for category, values in budget.items():
                ws3.append([category, values['amount'], values['percentage']])
            
            # Save to bytes
            from io import BytesIO
            output = BytesIO()
            wb.save(output)
            return output.getvalue()
            
        except Exception as e:
            logger.error(f"Excel creation failed: {e}")
            return None
    
    def _create_opportunity_email(self, opportunities: List[Dict], risk_reward: Dict, budget: Dict) -> str:
        """Create email digest for daily opportunities.
        
        Args:
            opportunities: List of opportunities
            risk_reward: Risk/reward matrix
            budget: Budget allocation
            
        Returns:
            HTML email content
        """
        try:
            # Create email content
            email_content = f"""
            <!DOCTYPE html>
            <html>
                <head>
                    <title>Daily Opportunity Report</title>
                    <style>
                        body {{ font-family: Arial, sans-serif; }}
                        table {{ border-collapse: collapse; width: 100%; margin-bottom: 20px; }}
                        th, td {{ border: 1px solid #ddd; padding: 8px; text-align: left; }}
                        th {{ background-color: #f5f5f5; }}
                        .summary {{ margin-bottom: 20px; }}
                    </style>
                </head>
                <body>
                    <h1>Daily Opportunity Report</h1>
                    
                    <div class="summary">
                        <h2>Top Opportunities</h2>
                        <table>
                            <tr>
                                <th>Item</th>
                                <th>Current Price</th>
                                <th>Estimated Value</th>
                                <th>Potential Profit</th>
                            </tr>
                            {''.join(f"<tr><td>{opp['item']}</td><td>${opp['current_price']}</td><td>${opp['estimated_value']}</td><td>${opp['potential_profit']}</td></tr>" for opp in opportunities[:5])}
                        </table>
                    </div>
                    
                    <div class="summary">
                        <h2>Risk/Reward Summary</h2>
                        <table>
                            <tr>
                                <th>Category</th>
                                <th>Risk Score</th>
                                <th>Reward Score</th>
                            </tr>
                            {''.join(f"<tr><td>{k}</td><td>{v['risk']}</td><td>{v['reward']}</td></tr>" for k, v in risk_reward.items())}
                        </table>
                    </div>
                    
                    <div class="summary">
                        <h2>Budget Allocation</h2>
                        <table>
                            <tr>
                                <th>Category</th>
                                <th>Amount</th>
                                <th>Percentage</th>
                            </tr>
                            {''.join(f"<tr><td>{k}</td><td>${v['amount']}</td><td>{v['percentage']}%</td></tr>" for k, v in budget.items())}
                        </table>
                    </div>
                    
                    <p>View the full report in your dashboard for more details.</p>
                </body>
            </html>
            """
            
            return email_content
            
        except Exception as e:
            logger.error(f"Email creation failed: {e}")
            return None 